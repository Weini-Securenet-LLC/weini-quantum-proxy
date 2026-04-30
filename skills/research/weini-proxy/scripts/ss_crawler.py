#!/usr/bin/env python3
"""
weini-proxy v5
GitHub 免费代理节点聚合器：抓取订阅、解析 ss/vmess/vless/trojan、兼容 Clash YAML、递归发现 provider URL、做 TCP 验证、Geo/ASN 丰富化、输出历史快照/差异/告警。

示例：
    PYTHONUNBUFFERED=1 python3 ss_crawler.py
    PYTHONUNBUFFERED=1 python3 ss_crawler.py --skip-verify --skip-enrich --max-nodes 100
    PYTHONUNBUFFERED=1 python3 ss_crawler.py --max-provider-depth 3 --history-dir /tmp/weini-history
    PYTHONUNBUFFERED=1 python3 ss_crawler.py --source-file references/default_sources.json
"""

from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import json
import re
import socket
import subprocess
import sys
import time
from collections import Counter, deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, unquote, urlparse

SKILL_DIR = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_FILE = SKILL_DIR / "references" / "default_sources.json"
DEFAULT_HISTORY_DIR = SKILL_DIR / "data" / "history"
SUPPORTED_PROTOCOLS = ("ss", "vmess", "vless", "trojan")
URI_PATTERN = re.compile(r"(?:ss|vmess|vless|trojan)://[^\s'\"<>]+")
BASE64_LINE_PATTERN = re.compile(r"^[A-Za-z0-9+/=]+$")
GENERIC_URL_PATTERN = re.compile(r"https?://[^\s'\"<>]+")
RAW_PROVIDER_HINT_PATTERN = re.compile(
    r"https?://[^\s'\"<>]+(?:\.ya?ml|\.txt|/sub(?:scribe)?[^\s'\"<>]*|/clash[^\s'\"<>]*|/mix[^\s'\"<>]*)",
    re.IGNORECASE,
)
IP_API_BATCH_URL = "http://ip-api.com/batch?fields=status,message,query,country,countryCode,regionName,city,isp,org,as,asname"


def b64_decode_pad(data: str) -> str:
    data = data.strip()
    if not data:
        return ""
    missing = len(data) % 4
    if missing:
        data += "=" * (4 - missing)
    try:
        return base64.b64decode(data, validate=False).decode("utf-8", errors="ignore")
    except Exception:
        return ""


def b64_encode_nopad(text: str) -> str:
    return base64.urlsafe_b64encode(text.encode("utf-8")).decode("utf-8").rstrip("=")


def ensure_yaml():
    try:
        import yaml  # type: ignore
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "PyYAML", "-q"])
        import yaml  # type: ignore
    return yaml


def read_sources(source_file: Path) -> list[dict]:
    raw = json.loads(source_file.read_text(encoding="utf-8"))
    sources = []
    for item in raw:
        repo = str(item.get("repo", "")).strip()
        if not repo:
            continue
        branches = item.get("branches") or ["main", "master"]
        paths = item.get("paths") or []
        include_readme = bool(item.get("include_readme", True))
        sources.append(
            {
                "repo": repo,
                "branches": [str(branch).strip() for branch in branches if str(branch).strip()],
                "paths": [str(path).strip() for path in paths if str(path).strip()],
                "include_readme": include_readme,
            }
        )
    return sources


def fetch_url(url: str, timeout: int = 30) -> str:
    try:
        result = subprocess.run(
            ["curl", "-sL", "--connect-timeout", str(timeout), "--max-time", str(timeout + 10), url],
            capture_output=True,
            text=True,
            timeout=timeout + 20,
        )
        return result.stdout
    except Exception:
        return ""


def post_json(url: str, payload: Any, timeout: int = 30) -> str:
    try:
        result = subprocess.run(
            [
                "curl",
                "-sL",
                "--connect-timeout",
                str(timeout),
                "--max-time",
                str(timeout + 10),
                "-H",
                "Content-Type: application/json",
                "-X",
                "POST",
                "-d",
                json.dumps(payload, ensure_ascii=False),
                url,
            ],
            capture_output=True,
            text=True,
            timeout=timeout + 20,
        )
        return result.stdout
    except Exception:
        return ""


def source_urls(source: dict) -> list[str]:
    urls = []
    repo = source["repo"]
    for branch in source["branches"]:
        for path in source["paths"]:
            urls.append(f"https://raw.githubusercontent.com/{repo}/{branch}/{path}")
        if source.get("include_readme", True):
            urls.append(f"https://raw.githubusercontent.com/{repo}/{branch}/README.md")
    return urls


def decode_content_variants(content: str) -> list[str]:
    variants = []
    seen = set()

    def add(value: str):
        value = value.strip()
        if value and value not in seen:
            seen.add(value)
            variants.append(value)

    add(content)
    stripped = content.strip()
    lines = stripped.splitlines()
    first_line = lines[0] if lines else ""
    collapsed = stripped.replace("\n", "")
    for candidate in (first_line, collapsed):
        decoded = b64_decode_pad(candidate)
        if decoded and any(f"{proto}://" in decoded for proto in SUPPORTED_PROTOCOLS):
            add(decoded)
    return variants


def split_host_port(hostport: str) -> tuple[str, int]:
    if hostport.startswith("["):
        end = hostport.index("]")
        host = hostport[1:end]
        port_str = hostport[end + 2 :]
    else:
        host, port_str = hostport.rsplit(":", 1)
    if not port_str.isdigit():
        raise ValueError("invalid port")
    port = int(port_str)
    if not host or not (1 <= port <= 65535):
        raise ValueError("invalid host/port")
    return host, port


def parse_ss_uri(uri: str) -> dict | None:
    try:
        body = uri[5:]
        name = ""
        if "#" in body:
            body, name = body.rsplit("#", 1)
            name = unquote(name)
        plugin = ""
        if "?" in body:
            body, query = body.split("?", 1)
            plugin = query
        if "@" in body:
            userinfo, hostport = body.split("@", 1)
            decoded_userinfo = b64_decode_pad(unquote(userinfo))
            if ":" not in decoded_userinfo:
                return None
            method, password = decoded_userinfo.split(":", 1)
        else:
            decoded = b64_decode_pad(body)
            if "@" not in decoded or ":" not in decoded:
                return None
            userinfo, hostport = decoded.rsplit("@", 1)
            method, password = userinfo.split(":", 1)
        host, port = split_host_port(hostport)
        return {
            "protocol": "ss",
            "name": name or "unnamed",
            "host": host,
            "port": port,
            "method": method,
            "credential": password,
            "network": "tcp",
            "tls": "unknown",
            "source_detail": plugin,
            "raw_uri": uri,
        }
    except Exception:
        return None


def parse_vmess_uri(uri: str) -> dict | None:
    try:
        payload = b64_decode_pad(uri[8:])
        if not payload:
            return None
        data = json.loads(payload)
        host = str(data.get("add") or "").strip()
        port = int(str(data.get("port") or "0"))
        if not host or not (1 <= port <= 65535):
            return None
        detail = {
            "path": data.get("path"),
            "host_header": data.get("host"),
            "sni": data.get("sni"),
            "aid": data.get("aid"),
            "alpn": data.get("alpn"),
        }
        return {
            "protocol": "vmess",
            "name": data.get("ps") or "unnamed",
            "host": host,
            "port": port,
            "method": data.get("scy") or data.get("type") or "auto",
            "credential": data.get("id") or "",
            "network": data.get("net") or "tcp",
            "tls": data.get("tls") or data.get("security") or "none",
            "source_detail": json.dumps(detail, ensure_ascii=False),
            "raw_uri": uri,
        }
    except Exception:
        return None


def parse_vless_or_trojan(uri: str, protocol: str) -> dict | None:
    try:
        parsed = urlparse(uri)
        host = str(parsed.hostname or "").strip()
        port = int(parsed.port or 0)
        if not host or not (1 <= port <= 65535):
            return None
        params = parse_qs(parsed.query)
        name = unquote(parsed.fragment) if parsed.fragment else "unnamed"
        network = params.get("type", [params.get("net", ["tcp"])[0]])[0]
        tls = params.get("security", ["tls" if protocol == "trojan" else "none"])[0]
        return {
            "protocol": protocol,
            "name": name,
            "host": host,
            "port": port,
            "method": network,
            "credential": unquote(parsed.username or ""),
            "network": network,
            "tls": tls,
            "source_detail": json.dumps({k: v[0] for k, v in params.items()}, ensure_ascii=False),
            "raw_uri": uri,
        }
    except Exception:
        return None


def parse_node(uri: str) -> dict | None:
    uri = uri.strip()
    if uri.startswith("ss://"):
        return parse_ss_uri(uri)
    if uri.startswith("vmess://"):
        return parse_vmess_uri(uri)
    if uri.startswith("vless://"):
        return parse_vless_or_trojan(uri, "vless")
    if uri.startswith("trojan://"):
        return parse_vless_or_trojan(uri, "trojan")
    return None


def node_to_ss_uri(proxy: dict) -> str | None:
    try:
        method = str(proxy.get("cipher") or proxy.get("method") or "").strip()
        password = str(proxy.get("password") or "").strip()
        server = str(proxy.get("server") or proxy.get("host") or "").strip()
        port = int(proxy.get("port") or 0)
        if not all([method, password, server]) or not (1 <= port <= 65535):
            return None
        name = quote(str(proxy.get("name") or "unnamed"))
        userinfo = b64_encode_nopad(f"{method}:{password}")
        return f"ss://{userinfo}@{server}:{port}#{name}"
    except Exception:
        return None


def node_to_vmess_uri(proxy: dict) -> str | None:
    try:
        server = str(proxy.get("server") or "").strip()
        port = int(proxy.get("port") or 0)
        uuid = str(proxy.get("uuid") or proxy.get("id") or "").strip()
        if not all([server, uuid]) or not (1 <= port <= 65535):
            return None
        payload = {
            "v": "2",
            "ps": str(proxy.get("name") or "unnamed"),
            "add": server,
            "port": str(port),
            "id": uuid,
            "aid": str(proxy.get("alterId") or proxy.get("alter-id") or 0),
            "scy": str(proxy.get("cipher") or proxy.get("method") or "auto"),
            "net": str(proxy.get("network") or proxy.get("net") or "tcp"),
            "type": str(proxy.get("type") or "none"),
            "host": str(proxy.get("host") or ""),
            "path": str(proxy.get("path") or ""),
            "tls": "tls" if proxy.get("tls") else "",
            "sni": str(proxy.get("servername") or proxy.get("serverName") or proxy.get("sni") or ""),
            "alpn": str(proxy.get("alpn") or ""),
        }
        encoded = base64.b64encode(json.dumps(payload, ensure_ascii=False).encode("utf-8")).decode("utf-8")
        return f"vmess://{encoded}"
    except Exception:
        return None


def node_to_vless_uri(proxy: dict) -> str | None:
    try:
        server = str(proxy.get("server") or "").strip()
        port = int(proxy.get("port") or 0)
        uuid = str(proxy.get("uuid") or proxy.get("id") or "").strip()
        if not all([server, uuid]) or not (1 <= port <= 65535):
            return None
        network = str(proxy.get("network") or proxy.get("net") or "tcp")
        security = "tls" if proxy.get("tls") else str(proxy.get("security") or "none")
        params = {"type": network, "security": security}
        if proxy.get("servername") or proxy.get("serverName") or proxy.get("sni"):
            params["sni"] = str(proxy.get("servername") or proxy.get("serverName") or proxy.get("sni"))
        if proxy.get("path"):
            params["path"] = str(proxy.get("path"))
        if proxy.get("alpn"):
            params["alpn"] = str(proxy.get("alpn"))
        query = "&".join(f"{quote(k)}={quote(str(v))}" for k, v in params.items() if str(v))
        name = quote(str(proxy.get("name") or "unnamed"))
        return f"vless://{uuid}@{server}:{port}?{query}#{name}"
    except Exception:
        return None


def node_to_trojan_uri(proxy: dict) -> str | None:
    try:
        server = str(proxy.get("server") or "").strip()
        port = int(proxy.get("port") or 0)
        password = str(proxy.get("password") or proxy.get("uuid") or "").strip()
        if not all([server, password]) or not (1 <= port <= 65535):
            return None
        network = str(proxy.get("network") or proxy.get("net") or "tcp")
        params = {"type": network, "security": "tls"}
        if proxy.get("servername") or proxy.get("serverName") or proxy.get("sni"):
            params["sni"] = str(proxy.get("servername") or proxy.get("serverName") or proxy.get("sni"))
        if proxy.get("path"):
            params["path"] = str(proxy.get("path"))
        if proxy.get("alpn"):
            params["alpn"] = str(proxy.get("alpn"))
        query = "&".join(f"{quote(k)}={quote(str(v))}" for k, v in params.items() if str(v))
        name = quote(str(proxy.get("name") or "unnamed"))
        return f"trojan://{password}@{server}:{port}?{query}#{name}"
    except Exception:
        return None


def convert_clash_proxy_to_uri(proxy: dict) -> str | None:
    proxy_type = str(proxy.get("type") or "").strip().lower()
    if proxy_type == "ss":
        return node_to_ss_uri(proxy)
    if proxy_type == "vmess":
        return node_to_vmess_uri(proxy)
    if proxy_type == "vless":
        return node_to_vless_uri(proxy)
    if proxy_type == "trojan":
        return node_to_trojan_uri(proxy)
    return None


def extract_clash_uris(text: str, protocols: set[str]) -> list[str]:
    yaml = ensure_yaml()
    extracted = []
    seen = set()
    try:
        documents = list(yaml.safe_load_all(text))
    except Exception:
        return []
    for doc in documents:
        if not isinstance(doc, dict):
            continue
        proxies = doc.get("proxies")
        if not isinstance(proxies, list):
            continue
        for proxy in proxies:
            if not isinstance(proxy, dict):
                continue
            uri = convert_clash_proxy_to_uri(proxy)
            if not uri:
                continue
            scheme = uri.split("://", 1)[0]
            if scheme in protocols and uri not in seen:
                seen.add(uri)
                extracted.append(uri)
    return extracted


def normalize_provider_url(url: str) -> str | None:
    url = url.strip().rstrip(")]}>,.;")
    if not url.startswith(("http://", "https://")):
        return None
    if any(token in url for token in ("telegram.me/", "t.me/", "github.com/")) and "raw.githubusercontent.com" not in url:
        return None
    if len(url) > 600:
        return None
    lowered = url.lower()
    interesting = (
        lowered.endswith((".yaml", ".yml", ".txt"))
        or "/sub" in lowered
        or "subscribe" in lowered
        or "clash" in lowered
        or "provider" in lowered
        or "mix" in lowered
    )
    return url if interesting else None


def extract_provider_urls(text: str) -> list[str]:
    yaml = ensure_yaml()
    urls = []
    seen = set()

    def add(url: str | None):
        normalized = normalize_provider_url(url or "")
        if normalized and normalized not in seen:
            seen.add(normalized)
            urls.append(normalized)

    try:
        documents = list(yaml.safe_load_all(text))
    except Exception:
        documents = []

    for doc in documents:
        if not isinstance(doc, dict):
            continue
        for providers_key in ("proxy-providers", "providers"):
            providers = doc.get(providers_key)
            if isinstance(providers, dict):
                for value in providers.values():
                    if isinstance(value, dict):
                        add(str(value.get("url") or ""))

    for match in RAW_PROVIDER_HINT_PATTERN.findall(text):
        add(match)
    for match in GENERIC_URL_PATTERN.findall(text):
        lowered = match.lower()
        if any(keyword in lowered for keyword in ("subscribe", "proxy-provider", "sub/", "sub_", "clash", "provider")):
            add(match)
    return urls


def extract_candidate_uris(text: str, protocols: set[str]) -> list[str]:
    candidates = []
    seen = set()

    def add_candidate(value: str):
        value = value.strip().rstrip(")]}>,.;")
        scheme = value.split("://", 1)[0]
        if scheme in protocols and value not in seen:
            seen.add(value)
            candidates.append(value)

    for match in URI_PATTERN.findall(text):
        add_candidate(match)

    for line in text.splitlines():
        candidate = line.strip()
        if len(candidate) > 40 and BASE64_LINE_PATTERN.match(candidate):
            decoded = b64_decode_pad(candidate)
            if decoded and any(f"{proto}://" in decoded for proto in protocols):
                for nested in URI_PATTERN.findall(decoded):
                    add_candidate(nested)

    for uri in extract_clash_uris(text, protocols):
        add_candidate(uri)
    return candidates


def initialize_node_defaults(node: dict) -> dict:
    node["reachable"] = None
    node["latency_ms"] = None
    node["fail_reason"] = ""
    node["resolved_ip"] = ""
    node["country"] = ""
    node["country_code"] = ""
    node["region"] = ""
    node["city"] = ""
    node["isp"] = ""
    node["org"] = ""
    node["asn"] = ""
    node["as_name"] = ""
    node["geo_status"] = ""
    node["provider_depth"] = int(node.get("provider_depth", 0))
    node["provider_chain"] = node.get("provider_chain", "")
    return node


def parsed_source_detail(node: dict) -> dict:
    detail = node.get("source_detail", "")
    if not detail:
        return {}
    try:
        parsed = json.loads(detail)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def node_fingerprint(node: dict) -> str:
    detail = parsed_source_detail(node)
    fingerprint_payload = {
        "protocol": node.get("protocol", ""),
        "host": node.get("host", ""),
        "port": node.get("port", ""),
        "credential": node.get("credential", ""),
        "network": node.get("network", ""),
        "tls": node.get("tls", ""),
        "method": node.get("method", ""),
        "sni": detail.get("sni") or detail.get("servername") or "",
        "path": detail.get("path") or "",
        "host_header": detail.get("host_header") or detail.get("host") or "",
        "alpn": detail.get("alpn") or "",
    }
    payload = json.dumps(fingerprint_payload, sort_keys=True, ensure_ascii=False)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def crawl_subscriptions(
    source_file: Path,
    protocols: set[str],
    timeout: int,
    max_nodes: int | None,
    provider_discovery: bool,
    max_discovered_urls: int,
    max_provider_depth: int,
) -> tuple[list[dict], list[dict]]:
    nodes = []
    source_stats = []
    seen_nodes = set()
    sources = read_sources(source_file)
    print(f"[信息] 载入源文件: {source_file} | 仓库数: {len(sources)}")

    for source in sources:
        print(f"[爬取] {source['repo']} ...")
        repo_new = 0
        repo_urls_checked = 0
        repo_discovered = 0
        repo_provider_urls_checked = 0
        repo_max_depth = 0
        pending_urls = deque((url, 0, "", source["repo"]) for url in source_urls(source))
        seen_urls = set()

        while pending_urls:
            url, depth, parent_url, discovered_from = pending_urls.popleft()
            if url in seen_urls:
                continue
            seen_urls.add(url)
            repo_urls_checked += 1
            repo_max_depth = max(repo_max_depth, depth)
            if depth > 0:
                repo_provider_urls_checked += 1
            content = fetch_url(url, timeout=timeout)
            if not content or len(content) < 16:
                continue

            if provider_discovery and depth < max_provider_depth and repo_discovered < max_discovered_urls:
                for discovered_url in extract_provider_urls(content):
                    if discovered_url in seen_urls:
                        continue
                    if any(item[0] == discovered_url for item in pending_urls):
                        continue
                    pending_urls.append((discovered_url, depth + 1, url, discovered_from))
                    repo_discovered += 1
                    if repo_discovered >= max_discovered_urls:
                        break

            for variant in decode_content_variants(content):
                for uri in extract_candidate_uris(variant, protocols):
                    node = parse_node(uri)
                    if not node:
                        continue
                    initialize_node_defaults(node)
                    node["source_repo"] = source["repo"]
                    node["source_url"] = url
                    node["provider_depth"] = depth
                    node["provider_chain"] = parent_url if parent_url else source["repo"]
                    fingerprint = node_fingerprint(node)
                    if fingerprint in seen_nodes:
                        continue
                    seen_nodes.add(fingerprint)
                    node["fingerprint"] = fingerprint
                    nodes.append(node)
                    repo_new += 1
                    if max_nodes and len(nodes) >= max_nodes:
                        source_stats.append(
                            {
                                "repo": source["repo"],
                                "urls_checked": repo_urls_checked,
                                "provider_urls_checked": repo_provider_urls_checked,
                                "new_nodes": repo_new,
                                "discovered_urls": repo_discovered,
                                "max_depth_reached": repo_max_depth,
                            }
                        )
                        print(f"  -> 已达到 max_nodes={max_nodes}，停止采集")
                        return nodes, source_stats

        source_stats.append(
            {
                "repo": source["repo"],
                "urls_checked": repo_urls_checked,
                "provider_urls_checked": repo_provider_urls_checked,
                "new_nodes": repo_new,
                "discovered_urls": repo_discovered,
                "max_depth_reached": repo_max_depth,
            }
        )
        print(
            f"  -> 当前累计节点: {len(nodes)} | 本仓新增: {repo_new} | 新发现URL: {repo_discovered} | 最大深度: {repo_max_depth}"
        )
    return nodes, source_stats


def resolve_first_ip(host: str, port: int) -> str:
    try:
        addrinfos = socket.getaddrinfo(host, port, type=socket.SOCK_STREAM)
    except Exception:
        return ""
    for _, _, _, _, sockaddr in addrinfos:
        if sockaddr:
            return str(sockaddr[0])
    return ""


def verify_tcp(node: dict, timeout: float = 3.0) -> dict:
    host = node["host"]
    port = node["port"]
    start = time.monotonic()
    try:
        addrinfos = socket.getaddrinfo(host, port, type=socket.SOCK_STREAM)
    except socket.gaierror:
        node["reachable"] = False
        node["fail_reason"] = "dns_fail"
        node["resolved_ip"] = ""
        return node
    except Exception as exc:
        node["reachable"] = False
        node["fail_reason"] = f"resolve_error:{str(exc)[:40]}"
        node["resolved_ip"] = ""
        return node

    last_error = "connect_fail"
    first_ip = ""
    for family, socktype, proto, _, sockaddr in addrinfos:
        sock = None
        try:
            if not first_ip and sockaddr:
                first_ip = str(sockaddr[0])
            sock = socket.socket(family, socktype, proto)
            sock.settimeout(timeout)
            result = sock.connect_ex(sockaddr)
            latency = round((time.monotonic() - start) * 1000)
            if result == 0:
                node["reachable"] = True
                node["latency_ms"] = latency
                node["fail_reason"] = ""
                node["resolved_ip"] = first_ip
                return node
            last_error = "refused" if result == 111 else f"errno_{result}"
        except socket.timeout:
            last_error = "timeout"
        except Exception as exc:
            last_error = str(exc)[:50]
        finally:
            if sock:
                try:
                    sock.close()
                except Exception:
                    pass

    node["reachable"] = False
    node["latency_ms"] = None
    node["fail_reason"] = last_error
    node["resolved_ip"] = first_ip
    return node


def verify_nodes(nodes: list[dict], timeout: float, workers: int) -> list[dict]:
    print(f"[验证] 开始 TCP 连通性测试 ({len(nodes)} 节点, {workers} 并发, {timeout}s 超时)")
    results = []
    total = len(nodes)
    done = 0
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(verify_tcp, dict(node), timeout): node for node in nodes}
        for future in as_completed(futures):
            done += 1
            results.append(future.result())
            if done % 50 == 0 or done == total:
                reachable = sum(1 for item in results if item.get("reachable"))
                print(f"  进度: {done}/{total} | 可用: {reachable}")
    return results


def resolve_missing_ips(nodes: list[dict], workers: int) -> None:
    unresolved = [node for node in nodes if not node.get("resolved_ip")]
    if not unresolved:
        return
    print(f"[丰富化] 开始解析 IP ({len(unresolved)} 节点)")

    def resolver(node: dict) -> tuple[str, str, int]:
        return resolve_first_ip(node["host"], node["port"]), node["host"], node["port"]

    mapping = {}
    with ThreadPoolExecutor(max_workers=max(1, min(workers, 32))) as pool:
        futures = {pool.submit(resolver, node): node for node in unresolved}
        for future in as_completed(futures):
            ip, host, port = future.result()
            mapping[(host, port)] = ip
    for node in unresolved:
        node["resolved_ip"] = mapping.get((node["host"], node["port"]), "")


def enrich_nodes_geo(nodes: list[dict], timeout: int = 12) -> None:
    ips = []
    seen = set()
    for node in nodes:
        ip = str(node.get("resolved_ip") or "").strip()
        if not ip or ip in seen:
            continue
        seen.add(ip)
        ips.append(ip)
    if not ips:
        return

    print(f"[丰富化] 开始 Geo/ASN 查询 ({len(ips)} 个 IP)")
    geo_map: dict[str, dict] = {}
    batch_size = 100
    for start in range(0, len(ips), batch_size):
        batch = [{"query": ip} for ip in ips[start : start + batch_size]]
        response = post_json(IP_API_BATCH_URL, batch, timeout=timeout)
        if not response:
            continue
        try:
            parsed = json.loads(response)
        except Exception:
            continue
        if not isinstance(parsed, list):
            continue
        for item in parsed:
            if not isinstance(item, dict):
                continue
            query = str(item.get("query") or "").strip()
            if query:
                geo_map[query] = item

    for node in nodes:
        ip = str(node.get("resolved_ip") or "").strip()
        if not ip:
            node["geo_status"] = node.get("geo_status") or "no_ip"
            continue
        item = geo_map.get(ip)
        if not item:
            node["geo_status"] = node.get("geo_status") or "lookup_failed"
            continue
        status = str(item.get("status") or "")
        node["geo_status"] = status or "unknown"
        if status != "success":
            continue
        node["country"] = str(item.get("country") or "")
        node["country_code"] = str(item.get("countryCode") or "")
        node["region"] = str(item.get("regionName") or "")
        node["city"] = str(item.get("city") or "")
        node["isp"] = str(item.get("isp") or "")
        node["org"] = str(item.get("org") or "")
        node["asn"] = str(item.get("as") or "")
        node["as_name"] = str(item.get("asname") or "")


def enrich_nodes(nodes: list[dict], workers: int, timeout: int) -> None:
    resolve_missing_ips(nodes, workers=workers)
    enrich_nodes_geo(nodes, timeout=timeout)


def protocol_stats(nodes: list[dict]) -> list[dict]:
    rows = []
    for protocol in SUPPORTED_PROTOCOLS:
        subset = [node for node in nodes if node["protocol"] == protocol]
        if not subset:
            continue
        reachable = [node for node in subset if node.get("reachable")]
        avg_latency = round(sum(node["latency_ms"] for node in reachable if node.get("latency_ms") is not None) / len(reachable), 1) if reachable else None
        rows.append(
            {
                "protocol": protocol,
                "total": len(subset),
                "reachable": len(reachable),
                "reachable_rate": round(len(reachable) * 100 / len(subset), 1),
                "avg_latency_ms": avg_latency,
            }
        )
    return rows


def top_counts(nodes: list[dict], field: str, top_n: int = 15) -> list[dict]:
    counter = Counter()
    for node in nodes:
        value = node.get(field)
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        counter[text] += 1
    return [{"name": name, "count": count} for name, count in counter.most_common(top_n)]


def stats_to_map(rows: list[dict]) -> dict[str, int]:
    return {str(row.get("name") or ""): int(row.get("count") or 0) for row in rows if str(row.get("name") or "")}


def compute_rank_delta(current_rows: list[dict], previous_rows: list[dict], top_n: int = 10) -> dict:
    current = stats_to_map(current_rows)
    previous = stats_to_map(previous_rows)
    names = set(current) | set(previous)
    changes = []
    for name in names:
        delta = current.get(name, 0) - previous.get(name, 0)
        if delta:
            changes.append({"name": name, "current": current.get(name, 0), "previous": previous.get(name, 0), "delta": delta})
    gains = sorted([item for item in changes if item["delta"] > 0], key=lambda item: (-item["delta"], item["name"]))[:top_n]
    losses = sorted([item for item in changes if item["delta"] < 0], key=lambda item: (item["delta"], item["name"]))[:top_n]
    return {"top_gains": gains, "top_losses": losses}


def history_key(node: dict) -> str:
    return str(node.get("fingerprint") or node_fingerprint(node))


def build_summary(nodes: list[dict], source_file: Path, source_stats: list[dict], previous_summary: dict | None = None) -> dict:
    reachable = [node for node in nodes if node.get("reachable")]
    fastest = sorted(reachable, key=lambda item: item.get("latency_ms", 10**9))[:20]
    current_keys = {history_key(node) for node in nodes}
    current_reachable_keys = {history_key(node) for node in reachable}

    previous_keys = set(previous_summary.get("node_keys", [])) if previous_summary else set()
    previous_reachable_keys = set(previous_summary.get("reachable_node_keys", [])) if previous_summary else set()

    new_nodes = len(current_keys - previous_keys)
    removed_nodes = len(previous_keys - current_keys)
    new_reachable = len(current_reachable_keys - previous_reachable_keys)
    removed_reachable = len(previous_reachable_keys - current_reachable_keys)

    country_rows = top_counts(nodes, "country")
    asn_rows = top_counts(nodes, "asn")
    isp_rows = top_counts(nodes, "isp")
    previous_country_rows = previous_summary.get("country_stats", []) if previous_summary else []
    previous_asn_rows = previous_summary.get("asn_stats", []) if previous_summary else []

    enriched_nodes = sum(1 for node in nodes if node.get("country") or node.get("asn"))
    return {
        "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "source_file": str(source_file),
        "total_nodes": len(nodes),
        "reachable_nodes": len(reachable),
        "reachable_rate": round(len(reachable) * 100 / len(nodes), 1) if nodes else 0.0,
        "enriched_nodes": enriched_nodes,
        "protocol_stats": protocol_stats(nodes),
        "source_stats": source_stats,
        "country_stats": country_rows,
        "asn_stats": asn_rows,
        "isp_stats": isp_rows,
        "provider_depth_stats": top_counts(nodes, "provider_depth"),
        "top_20_fastest": [
            {
                "protocol": node["protocol"],
                "name": node["name"],
                "host": node["host"],
                "port": node["port"],
                "latency_ms": node.get("latency_ms"),
                "country": node.get("country", ""),
                "asn": node.get("asn", ""),
                "source_repo": node.get("source_repo", ""),
                "provider_depth": node.get("provider_depth", 0),
            }
            for node in fastest
        ],
        "history_delta": {
            "has_previous_snapshot": previous_summary is not None,
            "new_nodes": new_nodes,
            "removed_nodes": removed_nodes,
            "new_reachable_nodes": new_reachable,
            "removed_reachable_nodes": removed_reachable,
            "previous_total_nodes": previous_summary.get("total_nodes") if previous_summary else None,
            "previous_reachable_nodes": previous_summary.get("reachable_nodes") if previous_summary else None,
        },
        "alert_delta": {
            "country": compute_rank_delta(country_rows, previous_country_rows),
            "asn": compute_rank_delta(asn_rows, previous_asn_rows),
        },
        "node_keys": sorted(current_keys),
        "reachable_node_keys": sorted(current_reachable_keys),
    }


def export_csv(nodes: list[dict], path: Path) -> None:
    headers = [
        "protocol",
        "name",
        "host",
        "port",
        "resolved_ip",
        "country",
        "country_code",
        "region",
        "city",
        "isp",
        "org",
        "asn",
        "as_name",
        "geo_status",
        "method",
        "credential",
        "network",
        "tls",
        "provider_depth",
        "provider_chain",
        "fingerprint",
        "reachable",
        "latency_ms",
        "fail_reason",
        "source_repo",
        "source_url",
        "source_detail",
        "raw_uri",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for node in nodes:
            writer.writerow({key: node.get(key, "") for key in headers})


def export_json(payload: dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def export_xlsx(nodes: list[dict], summary: dict, output_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    def format_header(ws, headers):
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        ws.freeze_panes = "A2"

    reachable = sorted([node for node in nodes if node.get("reachable")], key=lambda item: item.get("latency_ms", 10**9))

    ws1 = wb.active
    ws1.title = "可达节点"
    headers1 = ["#", "协议", "名称", "服务器", "端口", "IP", "国家", "ASN", "深度", "延迟(ms)", "来源仓库"]
    format_header(ws1, headers1)
    for idx, node in enumerate(reachable, 2):
        values = [idx - 1, node["protocol"], node["name"], node["host"], node["port"], node.get("resolved_ip", ""), node.get("country", ""), node.get("asn", ""), node.get("provider_depth", 0), node.get("latency_ms", ""), node.get("source_repo", "")]
        for col, value in enumerate(values, 1):
            cell = ws1.cell(row=idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        latency_cell = ws1.cell(row=idx, column=10)
        latency = node.get("latency_ms") or 9999
        if latency < 100:
            latency_cell.fill = green_fill
        elif latency < 200:
            latency_cell.fill = yellow_fill
        else:
            latency_cell.fill = red_fill
    ws1.auto_filter.ref = ws1.dimensions

    ws2 = wb.create_sheet("全部节点")
    headers2 = ["#", "协议", "名称", "服务器", "端口", "IP", "国家", "ASN", "方式", "网络", "TLS", "深度", "可达", "延迟(ms)", "失败原因", "来源仓库"]
    format_header(ws2, headers2)
    sorted_nodes = sorted(nodes, key=lambda item: (0 if item.get("reachable") else 1, item["protocol"], item.get("latency_ms") or 10**9))
    for idx, node in enumerate(sorted_nodes, 2):
        values = [idx - 1, node["protocol"], node["name"], node["host"], node["port"], node.get("resolved_ip", ""), node.get("country", ""), node.get("asn", ""), node["method"], node["network"], node["tls"], node.get("provider_depth", 0), "✅" if node.get("reachable") else "❌", node.get("latency_ms", ""), node.get("fail_reason", ""), node.get("source_repo", "")]
        for col, value in enumerate(values, 1):
            cell = ws2.cell(row=idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    ws2.auto_filter.ref = ws2.dimensions

    ws3 = wb.create_sheet("统计")
    rows = [
        ("生成时间", summary["generated_at_utc"]),
        ("总节点数", summary["total_nodes"]),
        ("可达节点数", summary["reachable_nodes"]),
        ("可达率", f"{summary['reachable_rate']}%"),
        ("已丰富化节点", summary["enriched_nodes"]),
        ("源文件", summary["source_file"]),
        ("新增节点", summary["history_delta"]["new_nodes"]),
        ("移除节点", summary["history_delta"]["removed_nodes"]),
        ("新增可达节点", summary["history_delta"]["new_reachable_nodes"]),
        ("移除可达节点", summary["history_delta"]["removed_reachable_nodes"]),
    ]
    for idx, (key, value) in enumerate(rows, 1):
        ws3.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws3.cell(row=idx, column=2, value=value)

    base_row = len(rows) + 3
    headers3 = ["协议", "总数", "可达", "可达率", "平均延迟(ms)"]
    for col, header in enumerate(headers3, 1):
        ws3.cell(row=base_row, column=col, value=header).font = Font(bold=True)
    for idx, row in enumerate(summary["protocol_stats"], base_row + 1):
        ws3.cell(row=idx, column=1, value=row["protocol"])
        ws3.cell(row=idx, column=2, value=row["total"])
        ws3.cell(row=idx, column=3, value=row["reachable"])
        ws3.cell(row=idx, column=4, value=f"{row['reachable_rate']}%")
        ws3.cell(row=idx, column=5, value=row["avg_latency_ms"])

    source_row = base_row + len(summary["protocol_stats"]) + 3
    source_headers = ["来源仓库", "检查URL数", "Provider URL数", "新增节点数", "发现URL数", "最大深度"]
    for col, header in enumerate(source_headers, 1):
        ws3.cell(row=source_row, column=col, value=header).font = Font(bold=True)
    for idx, row in enumerate(summary["source_stats"], source_row + 1):
        ws3.cell(row=idx, column=1, value=row["repo"])
        ws3.cell(row=idx, column=2, value=row["urls_checked"])
        ws3.cell(row=idx, column=3, value=row.get("provider_urls_checked", 0))
        ws3.cell(row=idx, column=4, value=row["new_nodes"])
        ws3.cell(row=idx, column=5, value=row.get("discovered_urls", 0))
        ws3.cell(row=idx, column=6, value=row.get("max_depth_reached", 0))

    ws4 = wb.create_sheet("地理分布")
    ws4.cell(row=1, column=1, value="国家").font = Font(bold=True)
    ws4.cell(row=1, column=2, value="数量").font = Font(bold=True)
    for idx, row in enumerate(summary.get("country_stats", []), 2):
        ws4.cell(row=idx, column=1, value=row["name"])
        ws4.cell(row=idx, column=2, value=row["count"])
    asn_row = max(2, len(summary.get("country_stats", [])) + 4)
    ws4.cell(row=asn_row, column=1, value="ASN").font = Font(bold=True)
    ws4.cell(row=asn_row, column=2, value="数量").font = Font(bold=True)
    for idx, row in enumerate(summary.get("asn_stats", []), asn_row + 1):
        ws4.cell(row=idx, column=1, value=row["name"])
        ws4.cell(row=idx, column=2, value=row["count"])

    ws5 = wb.create_sheet("变化告警")
    ws5.cell(row=1, column=1, value="维度").font = Font(bold=True)
    ws5.cell(row=1, column=2, value="名称").font = Font(bold=True)
    ws5.cell(row=1, column=3, value="当前").font = Font(bold=True)
    ws5.cell(row=1, column=4, value="上一轮").font = Font(bold=True)
    ws5.cell(row=1, column=5, value="变化").font = Font(bold=True)
    row_num = 2
    for dimension in ("country", "asn"):
        delta = summary.get("alert_delta", {}).get(dimension, {})
        for label, items in ((f"{dimension}_gain", delta.get("top_gains", [])), (f"{dimension}_loss", delta.get("top_losses", []))):
            for item in items:
                ws5.cell(row=row_num, column=1, value=label)
                ws5.cell(row=row_num, column=2, value=item["name"])
                ws5.cell(row=row_num, column=3, value=item["current"])
                ws5.cell(row=row_num, column=4, value=item["previous"])
                ws5.cell(row=row_num, column=5, value=item["delta"])
                row_num += 1

    ws6 = wb.create_sheet("原始URI")
    ws6.cell(row=1, column=1, value="raw_uri").font = Font(bold=True)
    for idx, node in enumerate(reachable, 2):
        ws6.cell(row=idx, column=1, value=node["raw_uri"])

    for ws in wb.worksheets:
        for letter, width in {"A": 18, "B": 20, "C": 30, "D": 26, "E": 12, "F": 18, "G": 16, "H": 24, "I": 14, "J": 18, "K": 16, "L": 18, "M": 18, "N": 20, "O": 24, "P": 24}.items():
            ws.column_dimensions[letter].width = width

    wb.save(output_path)


def parse_protocols(value: str) -> set[str]:
    protocols = {item.strip().lower() for item in value.split(",") if item.strip()}
    invalid = protocols.difference(SUPPORTED_PROTOCOLS)
    if invalid:
        raise argparse.ArgumentTypeError(f"unsupported protocols: {', '.join(sorted(invalid))}")
    return protocols


def derive_path(base: Path, suffix: str) -> Path:
    return base.with_suffix(suffix)


def load_previous_summary(history_dir: Path) -> dict | None:
    if not history_dir.exists():
        return None
    candidates = sorted(history_dir.glob("summary_*.json"))
    if not candidates:
        return None
    latest = candidates[-1]
    try:
        return json.loads(latest.read_text(encoding="utf-8"))
    except Exception:
        return None


def persist_history(summary: dict, nodes: list[dict], history_dir: Path) -> tuple[Path, Path]:
    history_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    summary_path = history_dir / f"summary_{ts}.json"
    nodes_path = history_dir / f"nodes_{ts}.json"
    export_json(summary, summary_path)
    export_json({"generated_at_utc": summary["generated_at_utc"], "nodes": nodes}, nodes_path)
    return summary_path, nodes_path


def build_delta_payload(summary: dict) -> dict:
    return {
        "generated_at_utc": summary["generated_at_utc"],
        "history_delta": summary["history_delta"],
        "current_total_nodes": summary["total_nodes"],
        "current_reachable_nodes": summary["reachable_nodes"],
        "enriched_nodes": summary.get("enriched_nodes", 0),
        "protocol_stats": summary["protocol_stats"],
        "country_stats": summary.get("country_stats", []),
        "asn_stats": summary.get("asn_stats", []),
        "alert_delta": summary.get("alert_delta", {}),
    }


def prune_summary(summary: dict) -> dict:
    trimmed = dict(summary)
    trimmed.pop("node_keys", None)
    trimmed.pop("reachable_node_keys", None)
    return trimmed


def main() -> int:
    parser = argparse.ArgumentParser(description="weini-proxy v5: 递归 provider 抓取 + Geo/ASN + 历史告警")
    parser.add_argument("--skip-verify", action="store_true", help="跳过 TCP 验证")
    parser.add_argument("--skip-enrich", action="store_true", help="跳过 Geo/ASN 丰富化")
    parser.add_argument("--skip-provider-discovery", action="store_true", help="跳过 provider URL 自动发现")
    parser.add_argument("--timeout", type=float, default=3.0, help="连接超时秒数")
    parser.add_argument("--workers", type=int, default=50, help="并发线程数")
    parser.add_argument("--output", type=str, default="", help="xlsx 输出路径")
    parser.add_argument("--json-output", type=str, default="", help="JSON 摘要输出路径")
    parser.add_argument("--csv-output", type=str, default="", help="CSV 明细输出路径")
    parser.add_argument("--delta-output", type=str, default="", help="历史差异 JSON 输出路径")
    parser.add_argument("--source-file", type=str, default=str(DEFAULT_SOURCE_FILE), help="JSON 源配置文件路径")
    parser.add_argument("--protocols", type=parse_protocols, default=set(SUPPORTED_PROTOCOLS), help="逗号分隔协议: ss,vmess,vless,trojan")
    parser.add_argument("--max-nodes", type=int, default=0, help="最多保留多少节点，0 表示不限")
    parser.add_argument("--history-dir", type=str, default=str(DEFAULT_HISTORY_DIR), help="历史快照目录")
    parser.add_argument("--no-history", action="store_true", help="不写入历史快照，也不读取上一份快照")
    parser.add_argument("--max-discovered-urls", type=int, default=20, help="每个仓库最多追加发现多少 provider URL")
    parser.add_argument("--max-provider-depth", type=int, default=2, help="provider URL 递归发现最大深度")
    args = parser.parse_args()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = Path(args.output) if args.output else Path(f"weini_proxy_{ts}.xlsx")
    json_output = Path(args.json_output) if args.json_output else derive_path(output, ".json")
    csv_output = Path(args.csv_output) if args.csv_output else derive_path(output, ".csv")
    delta_output = Path(args.delta_output) if args.delta_output else derive_path(output, ".delta.json")
    source_file = Path(args.source_file)
    history_dir = Path(args.history_dir)
    max_nodes = args.max_nodes or None

    if not source_file.exists():
        print(f"[错误] 源文件不存在: {source_file}")
        return 1

    print("=" * 64)
    print(" weini-proxy v5")
    print("=" * 64)
    print(
        f"[参数] protocols={','.join(sorted(args.protocols))} timeout={args.timeout}s workers={args.workers} "
        f"provider_discovery={'off' if args.skip_provider_discovery else 'on'} enrich={'off' if args.skip_enrich else 'on'} depth={args.max_provider_depth}"
    )

    previous_summary = None if args.no_history else load_previous_summary(history_dir)
    nodes, source_stats = crawl_subscriptions(
        source_file=source_file,
        protocols=args.protocols,
        timeout=max(int(args.timeout), 1),
        max_nodes=max_nodes,
        provider_discovery=not args.skip_provider_discovery,
        max_discovered_urls=max(0, args.max_discovered_urls),
        max_provider_depth=max(0, args.max_provider_depth),
    )
    print(f"[汇总] 共解析到 {len(nodes)} 个不重复节点")
    if not nodes:
        print("[警告] 未找到节点，退出")
        return 1

    if not args.skip_verify:
        nodes = verify_nodes(nodes, timeout=args.timeout, workers=args.workers)
    else:
        for node in nodes:
            node["reachable"] = False
            node["fail_reason"] = "skip_verify"

    if not args.skip_enrich:
        enrich_nodes(nodes, workers=args.workers, timeout=max(int(args.timeout * 4), 8))
    else:
        for node in nodes:
            node["geo_status"] = node.get("geo_status") or "skip_enrich"

    summary_full = build_summary(nodes, source_file=source_file, source_stats=source_stats, previous_summary=previous_summary)
    summary = prune_summary(summary_full)
    delta_payload = build_delta_payload(summary)

    export_xlsx(nodes, summary, output)
    export_csv(nodes, csv_output)
    export_json(summary, json_output)
    export_json(delta_payload, delta_output)

    history_summary_path = None
    history_nodes_path = None
    if not args.no_history:
        history_summary_path, history_nodes_path = persist_history(summary_full, nodes, history_dir)

    print(f"[完成] XLSX : {output}")
    print(f"[完成] CSV  : {csv_output}")
    print(f"[完成] JSON : {json_output}")
    print(f"[完成] DELTA: {delta_output}")
    if history_summary_path and history_nodes_path:
        print(f"[完成] HISTORY SUMMARY: {history_summary_path}")
        print(f"[完成] HISTORY NODES  : {history_nodes_path}")
    print(
        "[统计] 总节点: {total} | 可达: {reachable} | 可达率: {rate}% | 已丰富化: {enriched} | 新增: {new_nodes} | 移除: {removed}".format(
            total=summary["total_nodes"],
            reachable=summary["reachable_nodes"],
            rate=summary["reachable_rate"],
            enriched=summary.get("enriched_nodes", 0),
            new_nodes=summary["history_delta"]["new_nodes"],
            removed=summary["history_delta"]["removed_nodes"],
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
